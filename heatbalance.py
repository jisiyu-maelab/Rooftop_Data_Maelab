import shutil
import openpyxl


def heatbalance(templatepath, date_str):

    datapath = "G:\共有ドライブ\屋上Cadac2021\day" + "\\" + date_str + ".xlsx"
    new_file = "G:\共有ドライブ\屋上Cadac2021\day" + "\\熱収支" + date_str + ".xlsx"
    shutil.copy(templatepath, new_file)
    templatepath = new_file
    wb_temp = openpyxl.load_workbook(templatepath)
    wb_data = openpyxl.load_workbook(datapath)
    sheetname_data = wb_data.sheetnames  # 計測データの全シート名取得
    for sheetname in sheetname_data:
        for i, row in enumerate(wb_data[sheetname].iter_rows()):
            for j, cell in enumerate(row):
                wb_temp[sheetname].cell(row=i + 1, column=j + 1, value=cell.value)
    wb_temp.save(new_file)
