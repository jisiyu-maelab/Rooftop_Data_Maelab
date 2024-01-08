import numpy as np
# 異常値線形補間のログファイル作成
import pandas as pd
import openpyxl as px
from openpyxl.styles import Font

def errorcleaning(date_str, df_all):
    log_file = open(r"G:\共有ドライブ\屋上Cadac2021\day\log" + "\\" +  date_str + ".txt", "w")
    backup_file = r"G:\共有ドライブ\屋上Cadac2021\day\backup" + "\\" + date_str + ".xlsx"

    df_error = pd.DataFrame(index=df_all.index, columns=[])

    for column_name, item in df_all.iteritems():
        # 極端値の範囲定義
        if "温度" in column_name:
            low_limit = -20
            high_limit = 60
        elif "熱流計" in column_name:
            low_limit = -100
            high_limit = 100
        else:
            low_limit = -1000
            high_limit = 1500

        # 極端値→NaNでマーク
        ori_series = df_all[column_name].values.tolist()
        df_all[column_name] = df_all[column_name].mask(df_all[column_name] > high_limit, np.nan)
        df_all[column_name] = df_all[column_name].mask(df_all[column_name] < low_limit, np.nan)

        if df_all[column_name].isnull().sum() != 0:
            message = "nothing."
            if df_all[column_name].isnull().sum() == 8640:
                message = str(column_name) + " " + "断線です。"
                df_all[column_name] = 99999  #断線位置に99999を入力
            else:
                if df_all[column_name].isnull().sum() >= 2000:
                    message = str(column_name) + "は不安定です。" + str(df_all[column_name].isnull().sum()) + "個のデータが異常です。"
                elif df_all[column_name].isnull().sum() > 0:
                    message = str(column_name) + "の異常値を修正しました。"

                df_all[column_name].interpolate(limit_direction='both', inplace=True)  # 前後線形補間
                new_series = df_all[column_name].values.tolist()
                for i in range(len(ori_series)):
                    if ori_series[i] != new_series[i]:  # 書き換えがある場合
                        df_error[column_name] = np.nan
                        df_error[column_name][i] = ori_series[i]
                if message != "nothing.":
                    print(message)
                    log_file.writelines(message + "\r\n")
    df_error.to_excel(backup_file)  # backupデータ保存
    df_error_1min = df_error.resample("1T").mean()  # 一分間隔集計
    return df_error_1min


def red(date_str, df_error_1min):
    hb_file = r"G:\共有ドライブ\屋上Cadac2021\day" + "\\熱収支" + date_str + ".xlsx"
    wb_hb = px.load_workbook(hb_file)  # エクセル
    sheetname_hb = wb_hb.sheetnames  # 熱収支ファイル全シート名取得
    print(sheetname_hb)
    error_index = df_error_1min.index()
    print(error_index)
    for error_cl in error_index:
        for sheetname in sheetname_hb:  # エクセルのシート
            ws_hb = wb_hb[sheetname]
            df_hb = pd.read_excel(hb_file, sheet_name=sheetname)  # ヒートバランスシート
            if error_cl in df_hb.index():
                l_error_cl = df_error_1min[error_cl].to_list()
                row_l = []
                for i, j in enumerate(l_error_cl):
                    if j is None:
                        row_l.append(i)
                print(row_l)



                # print(error_index)
                # font = px.styles.fonts.Font(color='FF0000')  # メッソド
                #
                #         for i, row in enumerate(ws_hb.iter_rows()):
                #             for j, cell in enumerate(row):
                #                 if pd.isnull(ws_hb[]):
                #                     ws_hb[sheetname].cell(row=i + 1, column=j + 1)


    # df_hb =